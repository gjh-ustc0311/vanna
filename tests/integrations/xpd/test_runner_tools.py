import json
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from vanna.core.tool import ToolContext
from vanna.core.user import User
from vanna.integrations.local.agent_memory import DemoAgentMemory
from vanna.integrations.xpd.errors import XpdQueryTimeout
from vanna.integrations.xpd.files import XpdFileDeliveryError, XpdFileStore
from vanna.integrations.xpd.runner import XpdReadOnlyRunner
from vanna.integrations.xpd.sql_guard import XpdSqlGuard
from vanna.integrations.xpd.tools import (
    XPD_LLM_JSON_BYTE_LIMIT,
    RunXpdSqlArgs,
    RunXpdSqlTool,
    SearchXpdSchemaArgs,
    SearchXpdSchemaTool,
)


class QueryCursor:
    def __init__(self, rows, query_error=None):
        self.rows = list(rows)
        self.query_error = query_error
        self.description = None
        self.executed = []
        self.offset = 0
        self.fetch_sizes = []

    def execute(self, sql, params=None):
        self.executed.append((sql, params))
        if "_vanna_xpd_bounded_result" in sql:
            if self.query_error:
                raise self.query_error
            self.description = [("item_id",), ("pay_amt",)]

    def fetchmany(self, size):
        self.fetch_sizes.append(size)
        batch = self.rows[self.offset : self.offset + size]
        self.offset += len(batch)
        return batch

    def close(self):
        pass


class QueryConnection:
    def __init__(self, cursor):
        self.cursor_instance = cursor
        self.rolled_back = False

    def cursor(self):
        return self.cursor_instance

    def rollback(self):
        self.rolled_back = True

    def close(self):
        pass


class LoadedCatalog:
    def __init__(self, evidence):
        self.evidence = evidence


class FailingPublisher:
    def publish(self, artifact, store):
        del artifact, store
        raise XpdFileDeliveryError("safe failure")


@pytest.fixture
def tool_context():
    return ToolContext(
        user=User(id="u", group_memberships=["xpd"]),
        conversation_id="c",
        request_id="r",
        agent_memory=DemoAgentMemory(max_items=10),
    )


def make_runner(database_settings, schema_evidence, rows, tmp_path):
    cursor = QueryCursor(rows)
    connection = QueryConnection(cursor)
    store = XpdFileStore(tmp_path / "files")
    store.initialize()
    runner = XpdReadOnlyRunner(
        database_settings,
        XpdSqlGuard(schema_evidence),
        connection_factory=lambda: connection,
        file_store=store,
    )
    return runner, cursor, connection, store


@pytest.mark.asyncio
@pytest.mark.parametrize("row_count", [0, 1, 30, 31, 100, 101])
async def test_runner_boundary_collection(
    database_settings, schema_evidence, tmp_path, row_count
):
    first_item_id = 1_000_000_000_000
    rows = [
        (first_item_id + index, Decimal("1.20")) for index in range(row_count)
    ]
    runner, cursor, connection, store = make_runner(
        database_settings, schema_evidence, rows, tmp_path
    )

    result = await runner.run(
        "SELECT item_id, pay_amt FROM tb_live_goods_daily_stats", owner_id="owner"
    )

    assert result.returned_row_count == row_count
    assert len(result.analysis_rows) == min(row_count, 100)
    assert result.query_truncated is False
    assert (result.local_artifact is not None) is (row_count > 30)
    assert set(cursor.fetch_sizes) <= {500}
    assert connection.rolled_back is True
    if result.analysis_rows:
        assert result.analysis_rows[0]["item_id"] == first_item_id
    if result.local_artifact is not None:
        workbook = load_workbook(store.path_for(result.local_artifact), read_only=True)
        try:
            values = list(workbook[XLSX_SHEET_NAME].values)
        finally:
            workbook.close()
        assert values[0] == ("item_id", "pay_amt")
        assert len(values) == row_count + 1
        assert values[1][0] == str(first_item_id)
        assert values[1][1] == 1.2


XLSX_SHEET_NAME = "查询结果"


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("source_rows", "returned", "truncated"),
    [(20_000, 20_000, False), (20_001, 20_000, True)],
)
async def test_runner_caps_export_at_20000_and_uses_one_sentinel(
    database_settings,
    schema_evidence,
    tmp_path,
    source_rows,
    returned,
    truncated,
):
    rows = [(f"item-{index}", index) for index in range(source_rows)]
    runner, cursor, _connection, store = make_runner(
        database_settings, schema_evidence, rows, tmp_path
    )

    result = await runner.run(
        "SELECT item_id, pay_amt FROM tb_live_goods_daily_stats", owner_id="owner"
    )

    assert result.returned_row_count == returned
    assert result.query_truncated is truncated
    assert len(result.analysis_rows) == 100
    assert cursor.offset == source_rows
    assert result.local_artifact is not None
    assert result.local_artifact.row_count == 20_000
    assert result.local_artifact.truncated is truncated
    bounded = [
        sql for sql, _params in cursor.executed if "_vanna_xpd_bounded_result" in sql
    ]
    assert len(bounded) == 1
    assert bounded[0].endswith("LIMIT 20001")
    workbook = load_workbook(store.path_for(result.local_artifact), read_only=True)
    try:
        assert sum(1 for _row in workbook[XLSX_SHEET_NAME].values) == 20_001
    finally:
        workbook.close()


@pytest.mark.asyncio
async def test_runner_retries_only_connection(
    database_settings, schema_evidence, tmp_path
):
    cursor = QueryCursor([])
    connection = QueryConnection(cursor)
    attempts = 0

    def connect():
        nonlocal attempts
        attempts += 1
        if attempts == 1:
            raise OSError("unavailable")
        return connection

    store = XpdFileStore(tmp_path / "files")
    store.initialize()
    runner = XpdReadOnlyRunner(
        database_settings,
        XpdSqlGuard(schema_evidence),
        connection_factory=connect,
        file_store=store,
    )
    await runner.run("SELECT item_id, pay_amt FROM tb_live_goods_daily_stats")
    assert attempts == 2


@pytest.mark.asyncio
async def test_runner_does_not_replay_a_timed_out_query(
    database_settings, schema_evidence
):
    cursor = QueryCursor([], query_error=RuntimeError(3024, "raw database detail"))
    connection = QueryConnection(cursor)
    runner = XpdReadOnlyRunner(
        database_settings,
        XpdSqlGuard(schema_evidence),
        connection_factory=lambda: connection,
    )

    with pytest.raises(XpdQueryTimeout) as caught:
        await runner.run("SELECT item_id FROM tb_live_goods_daily_stats")

    assert "raw database detail" not in str(caught.value)
    assert (
        len([sql for sql, _ in cursor.executed if "_vanna_xpd_bounded_result" in sql])
        == 1
    )


@pytest.mark.asyncio
async def test_tools_require_same_turn_schema_and_use_30_row_preview(
    database_settings, schema_evidence, tool_context, tmp_path
):
    rows = [(f"item-{index}", index) for index in range(31)]
    runner, _cursor, _connection, store = make_runner(
        database_settings, schema_evidence, rows, tmp_path
    )
    run_tool = RunXpdSqlTool(runner, file_store=store)

    rejected = await run_tool.execute(
        tool_context,
        RunXpdSqlArgs(sql="SELECT item_id, pay_amt FROM tb_live_goods_daily_stats"),
    )
    assert rejected.success is False

    search = await SearchXpdSchemaTool(LoadedCatalog(schema_evidence)).execute(
        tool_context, SearchXpdSchemaArgs()
    )
    result = await run_tool.execute(
        tool_context,
        RunXpdSqlArgs(sql="SELECT item_id, pay_amt FROM tb_live_goods_daily_stats"),
    )

    assert search.success is True
    llm_result = json.loads(result.result_for_llm)
    assert len(llm_result["rows"]) == 31
    assert [component.type for component in result.components] == ["dataframe", "file"]
    assert len(result.components[0].rows) == 30
    assert result.components[1].url.startswith("/api/vanna/v3/files/")
    assert result.metadata["file_status"] == "available"
    assert "url" not in result.result_for_llm


@pytest.mark.asyncio
async def test_tool_enforces_llm_byte_budget_and_oss_failure_has_no_local_url(
    database_settings, schema_evidence, tool_context, tmp_path
):
    rows = [("x" * 10_000 + str(index), index) for index in range(31)]
    runner, _cursor, _connection, store = make_runner(
        database_settings, schema_evidence, rows, tmp_path
    )
    await SearchXpdSchemaTool(LoadedCatalog(schema_evidence)).execute(
        tool_context, SearchXpdSchemaArgs()
    )
    result = await RunXpdSqlTool(
        runner,
        file_store=store,
        oss_publisher=FailingPublisher(),  # type: ignore[arg-type]
    ).execute(
        tool_context,
        RunXpdSqlArgs(sql="SELECT item_id, pay_amt FROM tb_live_goods_daily_stats"),
    )

    payload = json.loads(result.result_for_llm)
    assert len(result.result_for_llm.encode("utf-8")) <= XPD_LLM_JSON_BYTE_LIMIT
    assert payload["rows_visible_to_llm"] < 31
    assert payload["file_status"] == "unavailable"
    assert "暂不可用" in payload["file_message"]
    assert [component.type for component in result.components] == ["dataframe"]
    assert result.metadata["exported_row_count"] == 31
    assert len(list((tmp_path / "files").glob("*/*/result.xlsx"))) == 1
