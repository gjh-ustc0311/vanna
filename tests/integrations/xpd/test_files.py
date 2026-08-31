import json
import os
import stat
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from types import SimpleNamespace

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from pydantic import SecretStr

from vanna.core.user import User
from vanna.core.user.request_context import RequestContext
from vanna.core.user.resolver import UserResolver
from vanna.integrations.xpd.config import XpdOssAccessSettings, XpdOssSettings
from vanna.integrations.xpd.files import (
    XLSX_MEDIA_TYPE,
    XpdFileExpired,
    XpdFileGenerationError,
    XpdFileNotFound,
    XpdFileStore,
    XpdOssPublisher,
    XpdXlsxWriter,
)
from vanna.servers.fastapi.xpd_files import register_xpd_file_routes


NOW = datetime(2026, 8, 31, 4, 5, 6, tzinfo=timezone.utc)


def committed_file(store, *, owner="owner", now=NOW, rows=31, truncated=False):
    draft = store.create_draft(owner)
    draft.staged_path.write_bytes(b"xlsx-content")
    os.chmod(draft.staged_path, 0o600)
    return store.commit(draft, row_count=rows, truncated=truncated)


def test_xlsx_writer_sanitizes_cells_and_preserves_supported_types(tmp_path):
    path = tmp_path / "result.xlsx"
    writer = XpdXlsxWriter(
        path,
        ["text", "amount", "when", "none", "bytes", "formula"],
    )
    writer.append(
        (
            "中文\x00值",
            Decimal("1.20"),
            datetime(2026, 8, 31, 12, 0, tzinfo=timezone.utc),
            None,
            b"abc",
            "  =SUM(A1:A2)",
        )
    )
    assert writer.finish() == 1

    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        row = list(workbook["查询结果"].values)[1]
    finally:
        workbook.close()
    assert row[0] == "中文值"
    assert row[1] == 1.2
    assert row[2] == datetime(2026, 8, 31, 20, 0)
    assert row[3] is None
    assert row[4] == "base64:YWJj"
    assert row[5] == "'  =SUM(A1:A2)"


@pytest.mark.parametrize("item_id_column", ["item_id", "商品ID"])
def test_xlsx_writer_exports_item_id_alias_as_exact_text(tmp_path, item_id_column):
    path = tmp_path / "result.xlsx"
    writer = XpdXlsxWriter(path, [item_id_column, "live_session_id", "count"])
    writer.append((601137401884, 42, 7))
    writer.append((9223372036854775807, 43, 8))
    writer.append((None, 44, 9))
    writer.append(("  =1+1", 45, 10))
    assert writer.finish() == 4

    workbook = load_workbook(path, data_only=False)
    try:
        sheet = workbook["查询结果"]
        assert sheet["A2"].value == "601137401884"
        assert sheet["A3"].value == "9223372036854775807"
        assert sheet["A4"].value is None
        assert sheet["A5"].value == "'  =1+1"
        assert sheet["A2"].data_type == "s"
        assert sheet["A3"].data_type == "s"
        assert sheet["A2"].number_format == "@"
        assert sheet["A3"].number_format == "@"
        assert sheet.column_dimensions["A"].width == 22
        assert sheet["B2"].value == 42
        assert sheet["B2"].data_type == "n"
        assert sheet["C2"].value == 7
        assert sheet["C2"].data_type == "n"
    finally:
        workbook.close()


def test_xlsx_writer_fails_safely_for_oversized_text(tmp_path):
    path = tmp_path / "result.xlsx"
    writer = XpdXlsxWriter(path, ["value"])
    with pytest.raises(XpdFileGenerationError):
        writer.append(("x" * 32_768,))
    writer.abort()
    assert not path.exists()


def test_file_store_is_owner_scoped_restart_safe_and_private(tmp_path):
    root = tmp_path / "files"
    store = XpdFileStore(root, clock=lambda: NOW)
    store.initialize()
    artifact = committed_file(store)

    assert artifact.name.startswith("xpd-query-20260831-120506-")
    assert store.resolve(artifact.file_id, "owner") == artifact
    with pytest.raises(XpdFileNotFound):
        store.resolve(artifact.file_id, "another-owner")
    restarted = XpdFileStore(root, clock=lambda: NOW)
    restarted.initialize()
    assert restarted.resolve(artifact.file_id, "owner") == artifact
    assert stat.S_IMODE(os.stat(store.path_for(artifact)).st_mode) == 0o600
    metadata_text = (store.path_for(artifact).parent / "metadata.json").read_text(
        encoding="utf-8"
    )
    assert "SELECT" not in metadata_text
    assert "question" not in metadata_text


def test_file_store_expiry_and_remote_delete_retry(tmp_path):
    store = XpdFileStore(tmp_path / "files", clock=lambda: NOW)
    store.initialize()
    artifact = committed_file(store)
    artifact = store.record_oss_receipt(artifact, "prefix/20260831/hash/file.xlsx")

    with pytest.raises(XpdFileExpired):
        store.resolve(
            artifact.file_id,
            artifact.owner_id,
            now=artifact.expires_at,
        )

    attempts = []

    def fail_once(object_key):
        attempts.append(object_key)
        raise RuntimeError("temporary")

    store.set_remote_deleter(fail_once)
    report = store.cleanup_expired(now=artifact.expires_at)
    assert report.remote_delete_retries == 1
    assert not store.path_for(artifact).exists()
    assert (store.path_for(artifact).parent / "metadata.json").exists()

    store.set_remote_deleter(lambda object_key: attempts.append(object_key))
    report = store.cleanup_expired(now=artifact.expires_at + timedelta(hours=1))
    assert report.removed == 1
    assert len(attempts) == 2


def test_file_store_rejects_symlinked_result(tmp_path):
    store = XpdFileStore(tmp_path / "files", clock=lambda: NOW)
    store.initialize()
    artifact = committed_file(store)
    result_path = store.path_for(artifact)
    result_path.unlink()
    result_path.symlink_to(tmp_path / "outside.xlsx")
    with pytest.raises(XpdFileNotFound):
        store.resolve(artifact.file_id, artifact.owner_id)


def test_file_store_rejects_overly_permissive_result(tmp_path):
    store = XpdFileStore(tmp_path / "files", clock=lambda: NOW)
    store.initialize()
    artifact = committed_file(store)
    os.chmod(store.path_for(artifact), 0o644)
    with pytest.raises(XpdFileNotFound):
        store.resolve(artifact.file_id, artifact.owner_id)


class HeaderUserResolver(UserResolver):
    async def resolve_user(self, request_context: RequestContext) -> User:
        return User(id=request_context.user_id or "anonymous")


def test_private_download_route_enforces_owner_and_headers(tmp_path):
    store = XpdFileStore(tmp_path / "files", clock=lambda: NOW)
    store.initialize()
    artifact = committed_file(store, owner="0")
    app = FastAPI()
    register_xpd_file_routes(app, store, HeaderUserResolver())

    with TestClient(app) as client:
        assert client.get(f"/api/vanna/v3/files/{artifact.file_id}").status_code == 422
        assert (
            client.get(
                f"/api/vanna/v3/files/{artifact.file_id}",
                headers=[("X-User-Id", "0"), ("X-User-Id", "0")],
            ).status_code
            == 422
        )
        assert (
            client.get(
                f"/api/vanna/v3/files/{artifact.file_id}",
                headers={"X-User-Id": "456"},
            ).status_code
            == 404
        )
        response = client.get(
            f"/api/vanna/v3/files/{artifact.file_id}",
            headers={"X-User-Id": "0"},
        )
        assert response.status_code == 200
        assert response.content == b"xlsx-content"
        assert response.headers["content-type"] == XLSX_MEDIA_TYPE
        assert response.headers["cache-control"] == "private, no-store"
        assert "attachment" in response.headers["content-disposition"]
        assert (
            client.get(
                "/api/vanna/v3/files/not-a-uuid",
                headers={"X-User-Id": "0"},
            ).status_code
            == 404
        )


def test_private_download_route_returns_410_at_expiry(tmp_path):
    clock = [NOW]
    store = XpdFileStore(tmp_path / "files", clock=lambda: clock[0])
    store.initialize()
    artifact = committed_file(store, owner="123")
    app = FastAPI()
    register_xpd_file_routes(app, store, HeaderUserResolver())

    with TestClient(app) as client:
        clock[0] = artifact.expires_at
        response = client.get(
            f"/api/vanna/v3/files/{artifact.file_id}",
            headers={"X-User-Id": "123"},
        )
        assert response.status_code == 410


def test_local_download_route_is_disabled_for_oss_delivery(tmp_path):
    store = XpdFileStore(tmp_path / "files", clock=lambda: NOW)
    store.initialize()
    app = FastAPI()
    register_xpd_file_routes(app, store, HeaderUserResolver(), enable_download=False)

    assert "/api/vanna/v3/files/{file_id}" not in {route.path for route in app.routes}


class FakeOssModule:
    class PutObjectRequest(SimpleNamespace):
        pass

    class GetObjectRequest(SimpleNamespace):
        pass

    class DeleteObjectRequest(SimpleNamespace):
        pass


class FakeOssClient:
    def __init__(self):
        self.uploads = []
        self.deletes = []
        self.presigns = []

    def put_object_from_file(self, request, path):
        self.uploads.append((request, path))
        return SimpleNamespace(status_code=200)

    def presign(self, request, expires):
        self.presigns.append((request, expires))
        return SimpleNamespace(
            url="https://bucket.oss.example.test/file.xlsx?signature=secret",
            expiration=NOW + expires,
        )

    def delete_object(self, request):
        self.deletes.append(request)
        return SimpleNamespace(status_code=204)


def test_oss_publisher_uploads_private_records_receipt_and_signs(tmp_path):
    store = XpdFileStore(tmp_path / "files", clock=lambda: NOW)
    store.initialize()
    artifact = committed_file(store)
    client = FakeOssClient()
    settings = XpdOssSettings(
        enabled=True,
        endpoint="https://oss.example.test",
        region="cn-test-1",
        bucket="test-bucket",
        prefix="private/query-files",
        access_key_id=SecretStr("id"),
        access_key_secret=SecretStr("secret"),
    )
    publisher = XpdOssPublisher(
        settings,
        XpdOssAccessSettings(url_ttl_seconds=86_400),
        client_factory=lambda: (FakeOssModule, client),
        clock=lambda: NOW,
    )

    publisher.initialize()
    published = publisher.publish(artifact, store)

    assert published.url.startswith("https://")
    assert published.expires_at == NOW + timedelta(days=1)
    request = client.uploads[0][0]
    assert request.acl == "private"
    assert request.forbid_overwrite is True
    assert request.content_type == XLSX_MEDIA_TYPE
    metadata = json.loads(
        (store.path_for(artifact).parent / "metadata.json").read_text(encoding="utf-8")
    )
    assert metadata["oss_object_key"].startswith("private/query-files/20260831/")
    assert "signature" not in json.dumps(metadata)
    publisher.delete(metadata["oss_object_key"])
    assert len(client.deletes) == 1
