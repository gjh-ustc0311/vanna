"""Private XPD XLSX storage, delivery, and lifecycle management."""

from __future__ import annotations

import hashlib
import json
import logging
import math
import os
import re
import shutil
import stat
import uuid
from dataclasses import dataclass, replace
from datetime import date, datetime, time, timedelta, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Mapping, Optional, Tuple
from urllib.parse import quote, urlsplit
from zoneinfo import ZoneInfo

from pydantic import (
    BaseModel,
    ConfigDict,
    Field,
    ValidationError,
    field_validator,
    model_validator,
)

from .config import XpdOssAccessSettings, XpdOssSettings
from .errors import XpdError


XPD_FILE_STORAGE_DIR = Path("datas/files")
XPD_FILE_TTL = timedelta(days=7)
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
XLSX_WORKSHEET_NAME = "查询结果"

_SHANGHAI = ZoneInfo("Asia/Shanghai")
_OWNER_DIRECTORY_PATTERN = re.compile(r"[0-9a-f]{16}")
_ILLEGAL_XLSX_CONTROLS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
_FORMULA_PREFIXES = ("=", "+", "-", "@")

logger = logging.getLogger(__name__)


class XpdFileNotFound(Exception):
    """The file does not exist or is not owned by the requester."""


class XpdFileExpired(Exception):
    """The requested file is no longer available."""


class XpdFileGenerationError(XpdError):
    """A safe boundary for XLSX or local persistence failures."""

    code = "xpd_file_generation_failed"
    public_message = "XPD result file could not be generated."


class XpdFileDeliveryError(XpdError):
    """A safe boundary for OSS upload or signing failures."""

    code = "xpd_file_delivery_failed"
    public_message = "XPD result file delivery is unavailable."


@dataclass(frozen=True)
class XpdFileDraft:
    file_id: uuid.UUID
    owner_id: str
    name: str
    staged_path: Path
    created_at: datetime
    expires_at: datetime


@dataclass(frozen=True)
class XpdFileArtifact:
    file_id: uuid.UUID
    owner_id: str
    name: str
    relative_path: str
    media_type: str
    size_bytes: int
    row_count: int
    truncated: bool
    created_at: datetime
    expires_at: datetime
    oss_object_key: Optional[str] = None


@dataclass(frozen=True)
class XpdPublishedFile:
    url: str
    expires_at: datetime


@dataclass(frozen=True)
class CleanupReport:
    removed: int = 0
    remote_delete_retries: int = 0
    invalid_entries: int = 0


class _StoredFileMetadata(BaseModel):
    model_config = ConfigDict(extra="forbid", strict=True)

    schema_version: int = Field(ge=1, le=1)
    file_id: uuid.UUID
    owner_id: str = Field(min_length=1, max_length=512)
    name: str = Field(min_length=1, max_length=255)
    relative_path: str = Field(min_length=1, max_length=1024)
    media_type: str
    size_bytes: int = Field(ge=0)
    row_count: int = Field(ge=0, le=20_000)
    truncated: bool
    created_at: datetime
    expires_at: datetime
    oss_object_key: Optional[str] = Field(default=None, max_length=1024)

    @field_validator("created_at", "expires_at")
    @classmethod
    def require_timezone(cls, value: datetime) -> datetime:
        if value.tzinfo is None or value.utcoffset() is None:
            raise ValueError("timestamps must include a timezone")
        return value

    @field_validator("name")
    @classmethod
    def validate_name(cls, value: str) -> str:
        if (
            value != value.strip()
            or any(character in value for character in {"/", "\\"})
            or any(ord(character) < 32 or ord(character) == 127 for character in value)
        ):
            raise ValueError("stored file name is invalid")
        return value

    @field_validator("media_type")
    @classmethod
    def validate_media_type(cls, value: str) -> str:
        if value != XLSX_MEDIA_TYPE:
            raise ValueError("stored file media type is invalid")
        return value

    @model_validator(mode="after")
    def validate_lifetime(self) -> "_StoredFileMetadata":
        if self.expires_at - self.created_at != XPD_FILE_TTL:
            raise ValueError("stored file lifetime is invalid")
        return self


def _utc_now() -> datetime:
    return datetime.now(timezone.utc)


def _owner_hash(owner_id: str) -> str:
    return hashlib.sha256(owner_id.encode("utf-8")).hexdigest()[:16]


def _fsync_directory(path: Path) -> None:
    try:
        descriptor = os.open(path, os.O_RDONLY)
    except OSError:
        return
    try:
        try:
            os.fsync(descriptor)
        except OSError:
            pass
    finally:
        os.close(descriptor)


def _safe_xlsx_text(value: Any) -> str:
    if isinstance(value, (Mapping, list, tuple, set)):
        text = json.dumps(value, ensure_ascii=False, default=str, separators=(",", ":"))
    else:
        text = str(value)
    text = _ILLEGAL_XLSX_CONTROLS.sub("", text)
    if len(text) > 32_767:
        raise XpdFileGenerationError(
            "XLSX cell text exceeds Excel's 32,767-character limit."
        )
    if text.lstrip().startswith(_FORMULA_PREFIXES):
        text = "'" + text
    return text


def _xlsx_cell_value(value: Any) -> Any:
    if value is None or isinstance(value, (bool, int, Decimal, date, time, timedelta)):
        if isinstance(value, datetime) and value.tzinfo is not None:
            return value.astimezone(_SHANGHAI).replace(tzinfo=None)
        if isinstance(value, time) and value.tzinfo is not None:
            return value.replace(tzinfo=None)
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else _safe_xlsx_text(value)
    if isinstance(value, memoryview):
        value = value.tobytes()
    if isinstance(value, bytes):
        import base64

        return "base64:" + base64.b64encode(value).decode("ascii")
    return _safe_xlsx_text(value)


class XpdXlsxWriter:
    """Incrementally creates one deterministic write-only XLSX workbook."""

    def __init__(self, path: Path, columns: Iterable[str]) -> None:
        try:
            from openpyxl import Workbook  # type: ignore[import-untyped]
            from openpyxl.cell import WriteOnlyCell  # type: ignore[import-untyped]
            from openpyxl.styles import (  # type: ignore[import-untyped]
                Alignment,
                Font,
                PatternFill,
            )
            from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
        except ImportError as exc:  # pragma: no cover - optional dependency boundary
            raise XpdFileGenerationError(
                "Install the 'xpd' extra to enable XLSX export."
            ) from exc

        self.path = path
        self.columns = list(columns)
        if not self.columns or len(self.columns) != len(set(self.columns)):
            raise XpdFileGenerationError("XLSX export requires unique columns.")
        headers = [_safe_xlsx_text(column) for column in self.columns]
        if len(headers) != len(set(headers)):
            raise XpdFileGenerationError(
                "XLSX columns are not unique after text sanitization."
            )
        self._write_only_cell = WriteOnlyCell
        self._row_count = 0
        self._closed = False
        self._workbook = Workbook(write_only=True)
        self._worksheet = self._workbook.create_sheet(XLSX_WORKSHEET_NAME)
        self._worksheet.freeze_panes = "A2"
        self._worksheet.sheet_view.showGridLines = False

        for index, header in enumerate(headers, start=1):
            width = 22 if "时间" in header else 14 if "日期" in header else 16
            self._worksheet.column_dimensions[get_column_letter(index)].width = width
        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        header_cells = []
        for header in headers:
            cell = WriteOnlyCell(self._worksheet, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            header_cells.append(cell)
        self._worksheet.append(header_cells)

    def append(self, raw_row: Any) -> None:
        if self._closed:
            raise XpdFileGenerationError("XLSX writer is already closed.")
        if isinstance(raw_row, Mapping):
            values = [raw_row.get(column) for column in self.columns]
        else:
            values = list(raw_row)
            if len(values) != len(self.columns):
                raise XpdFileGenerationError("XLSX row does not match its columns.")
        cells: List[Any] = []
        for value in values:
            normalized = _xlsx_cell_value(value)
            if isinstance(normalized, datetime):
                cell = self._write_only_cell(self._worksheet, value=normalized)
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
                cells.append(cell)
            elif isinstance(normalized, date):
                cell = self._write_only_cell(self._worksheet, value=normalized)
                cell.number_format = "yyyy-mm-dd"
                cells.append(cell)
            else:
                cells.append(normalized)
        self._worksheet.append(cells)
        self._row_count += 1

    def finish(self) -> int:
        if self._closed:
            raise XpdFileGenerationError("XLSX writer is already closed.")
        try:
            from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

            last_column = get_column_letter(len(self.columns))
            self._worksheet.auto_filter.ref = f"A1:{last_column}{self._row_count + 1}"
            self._workbook.save(self.path)
            return self._row_count
        except XpdFileGenerationError:
            raise
        except Exception as exc:
            raise XpdFileGenerationError("XLSX export could not be completed.") from exc
        finally:
            self._closed = True
            self._workbook.close()

    def abort(self) -> None:
        if not self._closed:
            self._closed = True
            try:
                self._worksheet.close()
            except Exception:
                pass
            try:
                self._workbook.close()
            except Exception:
                pass
        try:
            self.path.unlink()
        except FileNotFoundError:
            pass


class XpdFileStore:
    """Owner-scoped, restart-safe storage for generated query files."""

    def __init__(
        self,
        root: Path = XPD_FILE_STORAGE_DIR,
        *,
        clock: Callable[[], datetime] = _utc_now,
    ) -> None:
        self.root = Path(root)
        self._clock = clock
        self._staging = self.root / ".staging"
        self._remote_deleter: Optional[Callable[[str], None]] = None

    def set_remote_deleter(self, deleter: Callable[[str], None]) -> None:
        self._remote_deleter = deleter

    def initialize(self) -> None:
        if self.root.is_symlink():
            raise XpdFileGenerationError("XPD file storage must not be a symlink.")
        try:
            self.root.mkdir(parents=True, exist_ok=True, mode=0o700)
            self._staging.mkdir(exist_ok=True, mode=0o700)
            if self._staging.is_symlink():
                raise XpdFileGenerationError("XPD staging storage is invalid.")
            os.chmod(self.root, 0o700)
            os.chmod(self._staging, 0o700)
            probe = self._staging / f".probe-{uuid.uuid4().hex}"
            descriptor = os.open(probe, os.O_WRONLY | os.O_CREAT | os.O_EXCL, 0o600)
            os.close(descriptor)
            probe.unlink()
        except XpdFileGenerationError:
            raise
        except OSError as exc:
            raise XpdFileGenerationError("XPD file storage is not writable.") from exc

    def create_draft(self, owner_id: str) -> XpdFileDraft:
        if not owner_id:
            raise XpdFileGenerationError("XPD file owner is required.")
        file_id = uuid.uuid4()
        created_at = self._clock().astimezone(timezone.utc)
        name = (
            f"xpd-query-{created_at.astimezone(_SHANGHAI):%Y%m%d-%H%M%S}-"
            f"{file_id.hex[:8]}.xlsx"
        )
        return XpdFileDraft(
            file_id=file_id,
            owner_id=owner_id,
            name=name,
            staged_path=self._staging / f"{file_id}.xlsx",
            created_at=created_at,
            expires_at=created_at + XPD_FILE_TTL,
        )

    def discard(self, draft: XpdFileDraft) -> None:
        try:
            draft.staged_path.unlink()
        except FileNotFoundError:
            pass

    def commit(
        self,
        draft: XpdFileDraft,
        *,
        row_count: int,
        truncated: bool,
    ) -> XpdFileArtifact:
        try:
            staged_stat = os.lstat(draft.staged_path)
        except OSError as exc:
            raise XpdFileGenerationError(
                "The staged XLSX file is unavailable."
            ) from exc
        if stat.S_ISLNK(staged_stat.st_mode) or not stat.S_ISREG(staged_stat.st_mode):
            raise XpdFileGenerationError("The staged XLSX file is invalid.")

        owner_directory = _owner_hash(draft.owner_id)
        owner_path = self.root / owner_directory
        final_directory = owner_path / str(draft.file_id)
        commit_directory = self._staging / f"commit-{draft.file_id}"
        relative_path = f"{owner_directory}/{draft.file_id}/result.xlsx"
        artifact = XpdFileArtifact(
            file_id=draft.file_id,
            owner_id=draft.owner_id,
            name=draft.name,
            relative_path=relative_path,
            media_type=XLSX_MEDIA_TYPE,
            size_bytes=staged_stat.st_size,
            row_count=row_count,
            truncated=truncated,
            created_at=draft.created_at,
            expires_at=draft.expires_at,
        )
        try:
            if owner_path.is_symlink():
                raise XpdFileGenerationError("The owner storage directory is invalid.")
            owner_path.mkdir(exist_ok=True, mode=0o700)
            os.chmod(owner_path, 0o700)
            commit_directory.mkdir(mode=0o700)
            result_path = commit_directory / "result.xlsx"
            os.replace(draft.staged_path, result_path)
            os.chmod(result_path, 0o600)
            with result_path.open("rb") as result_file:
                os.fsync(result_file.fileno())
            self._write_metadata(commit_directory / "metadata.json", artifact)
            os.replace(commit_directory, final_directory)
            _fsync_directory(owner_path)
            return artifact
        except FileExistsError as exc:
            raise XpdFileGenerationError(
                "The generated file identifier collided."
            ) from exc
        except Exception as exc:
            if isinstance(exc, XpdFileGenerationError):
                raise
            raise XpdFileGenerationError(
                "The generated XLSX could not be stored."
            ) from exc
        finally:
            if commit_directory.exists() and not commit_directory.is_symlink():
                shutil.rmtree(commit_directory, ignore_errors=True)
            self.discard(draft)

    def _write_metadata(self, path: Path, artifact: XpdFileArtifact) -> None:
        metadata = _StoredFileMetadata(
            schema_version=1,
            file_id=artifact.file_id,
            owner_id=artifact.owner_id,
            name=artifact.name,
            relative_path=artifact.relative_path,
            media_type=artifact.media_type,
            size_bytes=artifact.size_bytes,
            row_count=artifact.row_count,
            truncated=artifact.truncated,
            created_at=artifact.created_at,
            expires_at=artifact.expires_at,
            oss_object_key=artifact.oss_object_key,
        )
        temporary = path.with_name(f".{path.name}.tmp-{uuid.uuid4().hex}")
        payload = metadata.model_dump_json().encode("utf-8")
        descriptor = os.open(temporary, os.O_WRONLY | os.O_CREAT | os.O_EXCL, 0o600)
        try:
            view = memoryview(payload)
            while view:
                written = os.write(descriptor, view)
                if written <= 0:
                    raise OSError("metadata write did not progress")
                view = view[written:]
            os.fsync(descriptor)
        finally:
            os.close(descriptor)
        try:
            os.replace(temporary, path)
            os.chmod(path, 0o600)
            _fsync_directory(path.parent)
        finally:
            try:
                temporary.unlink()
            except FileNotFoundError:
                pass

    def _artifact_from_metadata(self, metadata: _StoredFileMetadata) -> XpdFileArtifact:
        return XpdFileArtifact(
            file_id=metadata.file_id,
            owner_id=metadata.owner_id,
            name=metadata.name,
            relative_path=metadata.relative_path,
            media_type=metadata.media_type,
            size_bytes=metadata.size_bytes,
            row_count=metadata.row_count,
            truncated=metadata.truncated,
            created_at=metadata.created_at,
            expires_at=metadata.expires_at,
            oss_object_key=metadata.oss_object_key,
        )

    def _load_owned(
        self, file_id: uuid.UUID, owner_id: str
    ) -> Tuple[XpdFileArtifact, Path, Path]:
        owner_directory = _owner_hash(owner_id)
        owner_path = self.root / owner_directory
        directory = owner_path / str(file_id)
        metadata_path = directory / "metadata.json"
        result_path = directory / "result.xlsx"
        try:
            owner_stat = os.lstat(owner_path)
            directory_stat = os.lstat(directory)
            metadata_stat = os.lstat(metadata_path)
        except OSError as exc:
            raise XpdFileNotFound() from exc
        if (
            stat.S_ISLNK(owner_stat.st_mode)
            or not stat.S_ISDIR(owner_stat.st_mode)
            or owner_stat.st_mode & 0o077
            or stat.S_ISLNK(directory_stat.st_mode)
            or not stat.S_ISDIR(directory_stat.st_mode)
            or directory_stat.st_mode & 0o077
            or stat.S_ISLNK(metadata_stat.st_mode)
            or not stat.S_ISREG(metadata_stat.st_mode)
            or metadata_stat.st_mode & 0o077
        ):
            raise XpdFileNotFound()
        try:
            metadata = _StoredFileMetadata.model_validate_json(
                metadata_path.read_text(encoding="utf-8")
            )
        except (OSError, UnicodeError, ValidationError) as exc:
            raise XpdFileNotFound() from exc
        expected_relative = f"{owner_directory}/{file_id}/result.xlsx"
        if (
            metadata.file_id != file_id
            or metadata.owner_id != owner_id
            or metadata.relative_path != expected_relative
        ):
            raise XpdFileNotFound()
        return self._artifact_from_metadata(metadata), metadata_path, result_path

    def resolve(
        self,
        file_id: uuid.UUID,
        owner_id: str,
        *,
        now: Optional[datetime] = None,
    ) -> XpdFileArtifact:
        artifact, _metadata_path, result_path = self._load_owned(file_id, owner_id)
        current = (now or self._clock()).astimezone(timezone.utc)
        if current >= artifact.expires_at:
            raise XpdFileExpired()
        try:
            result_stat = os.lstat(result_path)
        except OSError as exc:
            raise XpdFileNotFound() from exc
        if (
            stat.S_ISLNK(result_stat.st_mode)
            or not stat.S_ISREG(result_stat.st_mode)
            or result_stat.st_mode & 0o077
        ):
            raise XpdFileNotFound()
        if result_stat.st_size != artifact.size_bytes:
            raise XpdFileNotFound()
        return artifact

    def path_for(self, artifact: XpdFileArtifact) -> Path:
        expected = (
            self.root
            / _owner_hash(artifact.owner_id)
            / str(artifact.file_id)
            / "result.xlsx"
        )
        if artifact.relative_path != str(expected.relative_to(self.root)):
            raise XpdFileNotFound()
        return expected

    def record_oss_receipt(
        self, artifact: XpdFileArtifact, object_key: str
    ) -> XpdFileArtifact:
        current, metadata_path, _result_path = self._load_owned(
            artifact.file_id, artifact.owner_id
        )
        if current != artifact or not object_key:
            raise XpdFileDeliveryError("OSS receipt metadata is inconsistent.")
        updated = replace(current, oss_object_key=object_key)
        self._write_metadata(metadata_path, updated)
        return updated

    def cleanup_expired(self, *, now: Optional[datetime] = None) -> CleanupReport:
        current = (now or self._clock()).astimezone(timezone.utc)
        removed = retries = invalid = 0
        if not self.root.exists():
            return CleanupReport()
        for owner_path in self.root.iterdir():
            if owner_path.name == ".staging":
                continue
            if (
                owner_path.is_symlink()
                or not owner_path.is_dir()
                or not _OWNER_DIRECTORY_PATTERN.fullmatch(owner_path.name)
            ):
                invalid += 1
                continue
            for directory in owner_path.iterdir():
                try:
                    file_id = uuid.UUID(directory.name)
                except ValueError:
                    invalid += 1
                    continue
                if directory.is_symlink() or not directory.is_dir():
                    invalid += 1
                    continue
                metadata_path = directory / "metadata.json"
                try:
                    metadata = _StoredFileMetadata.model_validate_json(
                        metadata_path.read_text(encoding="utf-8")
                    )
                except (OSError, UnicodeError, ValidationError):
                    invalid += 1
                    continue
                if metadata.file_id != file_id or current < metadata.expires_at:
                    continue
                result_path = directory / "result.xlsx"
                try:
                    if result_path.exists() and not result_path.is_symlink():
                        result_path.unlink()
                except OSError:
                    logger.warning("Expired XPD file local deletion failed")
                if metadata.oss_object_key and self._remote_deleter is not None:
                    try:
                        self._remote_deleter(metadata.oss_object_key)
                    except Exception:
                        retries += 1
                        logger.warning(
                            "Expired XPD file remote deletion will be retried"
                        )
                        continue
                try:
                    metadata_path.unlink()
                    directory.rmdir()
                    removed += 1
                except OSError:
                    logger.warning("Expired XPD file metadata deletion failed")
            try:
                owner_path.rmdir()
            except OSError:
                pass
        return CleanupReport(
            removed=removed,
            remote_delete_retries=retries,
            invalid_entries=invalid,
        )


class XpdOssPublisher:
    """Uploads private XLSX files and produces short-lived HTTPS URLs."""

    def __init__(
        self,
        settings: XpdOssSettings,
        access: XpdOssAccessSettings,
        *,
        client_factory: Optional[Callable[[], Tuple[Any, Any]]] = None,
        clock: Callable[[], datetime] = _utc_now,
    ) -> None:
        if not settings.enabled:
            raise ValueError("OSS publisher requires enabled OSS settings")
        self.settings = settings
        self.access = access
        self._client_factory = client_factory
        self._clock = clock

    def _client(self) -> Tuple[Any, Any]:
        if self._client_factory is not None:
            return self._client_factory()
        try:
            import alibabacloud_oss_v2 as oss  # type: ignore[import-untyped]
        except ImportError as exc:  # pragma: no cover - optional dependency boundary
            raise XpdFileDeliveryError(
                "Install the 'xpd' extra to enable OSS delivery."
            ) from exc
        access_key_id = self.settings.access_key_id
        access_key_secret = self.settings.access_key_secret
        if access_key_id is None or access_key_secret is None:
            raise XpdFileDeliveryError("OSS credentials are unavailable.")
        credentials = oss.credentials.StaticCredentialsProvider(
            access_key_id.get_secret_value(),
            access_key_secret.get_secret_value(),
            (
                self.settings.security_token.get_secret_value()
                if self.settings.security_token is not None
                else None
            ),
        )
        config = oss.config.load_default()
        config.credentials_provider = credentials
        config.region = self.settings.region
        config.endpoint = self.settings.endpoint
        config.connect_timeout = 10
        config.readwrite_timeout = 30
        config.retry_max_attempts = 2
        return oss, oss.Client(config)

    def initialize(self) -> None:
        try:
            oss, client = self._client()
            request_types = (
                getattr(oss, "PutObjectRequest", None),
                getattr(oss, "GetObjectRequest", None),
                getattr(oss, "DeleteObjectRequest", None),
            )
            if any(request_type is None for request_type in request_types):
                raise RuntimeError("OSS request models are unavailable")
            if (
                not callable(getattr(client, "presign", None))
                or not callable(getattr(client, "delete_object", None))
                or not (
                    callable(getattr(client, "put_object_from_file", None))
                    or callable(getattr(client, "put_object", None))
                )
            ):
                raise RuntimeError("OSS client methods are unavailable")
        except XpdFileDeliveryError:
            raise
        except Exception as exc:
            raise XpdFileDeliveryError("OSS client initialization failed.") from exc

    def _object_key(self, artifact: XpdFileArtifact) -> str:
        created = artifact.created_at.astimezone(_SHANGHAI)
        return (
            f"{self.settings.prefix}/{created:%Y%m%d}/"
            f"{_owner_hash(artifact.owner_id)}/{artifact.file_id}.xlsx"
        )

    def publish(
        self, artifact: XpdFileArtifact, store: XpdFileStore
    ) -> XpdPublishedFile:
        current = store.resolve(artifact.file_id, artifact.owner_id)
        if current != artifact:
            raise XpdFileDeliveryError("Local file metadata is inconsistent.")
        object_key = self._object_key(artifact)
        source_path = store.path_for(artifact)
        disposition = f"attachment; filename*=UTF-8''{quote(artifact.name, safe='')}"
        try:
            oss, client = self._client()
            request = oss.PutObjectRequest(
                bucket=self.settings.bucket,
                key=object_key,
                acl="private",
                content_type=artifact.media_type,
                content_disposition=disposition,
                content_length=artifact.size_bytes,
                forbid_overwrite=True,
            )
            put_from_file = getattr(client, "put_object_from_file", None)
            if callable(put_from_file):
                response = put_from_file(request, str(source_path))
            else:  # pragma: no cover - old/minimal client compatibility
                request.body = source_path.read_bytes()
                response = client.put_object(request)
            status = int(getattr(response, "status_code", 0) or 0)
            if not 200 <= status < 300:
                raise RuntimeError("OSS upload returned an unsuccessful status")

            try:
                store.record_oss_receipt(artifact, object_key)
            except Exception:
                try:
                    client.delete_object(
                        oss.DeleteObjectRequest(
                            bucket=self.settings.bucket, key=object_key
                        )
                    )
                except Exception:
                    logger.warning("OSS compensation deletion failed")
                raise

            signed = client.presign(
                oss.GetObjectRequest(
                    bucket=self.settings.bucket,
                    key=object_key,
                    response_content_disposition=disposition,
                ),
                expires=timedelta(seconds=self.access.url_ttl_seconds),
            )
            url = str(getattr(signed, "url", "") or "")
            parsed = urlsplit(url)
            if (
                len(url) > 8192
                or url != url.strip()
                or parsed.scheme != "https"
                or not parsed.netloc
                or parsed.username is not None
                or parsed.password is not None
                or parsed.fragment
            ):
                raise RuntimeError("OSS returned an invalid signed URL")
            expiration = getattr(signed, "expiration", None)
            if not isinstance(expiration, datetime):
                expiration = self._clock() + timedelta(
                    seconds=self.access.url_ttl_seconds
                )
            elif expiration.tzinfo is None:
                expiration = expiration.replace(tzinfo=timezone.utc)
            return XpdPublishedFile(url=url, expires_at=expiration)
        except Exception as exc:
            if isinstance(exc, XpdFileDeliveryError):
                raise
            raise XpdFileDeliveryError(
                "XPD file delivery is temporarily unavailable."
            ) from exc

    def delete(self, object_key: str) -> None:
        expected_prefix = f"{self.settings.prefix}/"
        if (
            not object_key.startswith(expected_prefix)
            or not object_key.endswith(".xlsx")
            or "\\" in object_key
            or any(part in {"", ".", ".."} for part in object_key.split("/"))
        ):
            raise XpdFileDeliveryError("OSS object receipt is invalid.")
        try:
            oss, client = self._client()
            response = client.delete_object(
                oss.DeleteObjectRequest(bucket=self.settings.bucket, key=object_key)
            )
            status = int(getattr(response, "status_code", 0) or 0)
            if not 200 <= status < 300:
                raise RuntimeError("OSS deletion returned an unsuccessful status")
        except Exception as exc:
            raise XpdFileDeliveryError("OSS deletion failed.") from exc


__all__ = [
    "CleanupReport",
    "XLSX_MEDIA_TYPE",
    "XPD_FILE_STORAGE_DIR",
    "XpdFileArtifact",
    "XpdFileDeliveryError",
    "XpdFileDraft",
    "XpdFileExpired",
    "XpdFileGenerationError",
    "XpdFileNotFound",
    "XpdFileStore",
    "XpdOssPublisher",
    "XpdPublishedFile",
    "XpdXlsxWriter",
]
